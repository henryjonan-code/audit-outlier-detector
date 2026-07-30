"""Rate Change Anomaly Detector — Hotel FO Audit"""

import re
from pathlib import Path


def parse_amount(v):
    if v is None:
        return None
    s = str(v).replace(',', '').replace("'", '').strip()
    try:
        return float(s)
    except Exception:
        return None


def is_afterhours(time_str):
    if not time_str:
        return False
    try:
        h = int(str(time_str).split(':')[0])
        return h >= 22 or h < 6
    except Exception:
        return False


def load_file(path):
    path = Path(path)
    ext = path.suffix.lower()

    if ext == '.numbers':
        from numbers_parser import Document
        doc = Document(str(path))
        sheet = doc.sheets[0]
        table = sheet.tables[0]
        rows = []
        for i, row in enumerate(table.iter_rows()):
            if i < 3:
                continue
            vals = [cell.value for cell in row]
            if vals[0] is None:
                continue
            rows.append(vals)
        return rows

    elif ext in ('.xlsx', '.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        rows = []
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row[0] and 'reservation' in str(row[0]).lower():
                    header_found = True
                continue
            if row[0] is None:
                continue
            rows.append(list(row))
        return rows

    elif ext == '.csv':
        import csv
        rows = []
        with open(str(path), encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            header_found = False
            for row in reader:
                if not header_found:
                    if row and 'reservation' in row[0].lower():
                        header_found = True
                    continue
                if not row or not row[0]:
                    continue
                rows.append(row)
        return rows

    raise ValueError(f"Format tidak didukung: {ext}")


def analyze(rows):
    # Sheet 1: Rate turun > 30%
    s1, seen1 = [], set()
    for r in rows:
        r1 = parse_amount(r[9])
        r2 = parse_amount(r[10])
        v = parse_amount(r[11])
        if not (r1 and r2 and r2 > 0 and v and v < 0):
            continue
        pct = abs(v) / r2 * 100
        if pct < 30:
            continue
        key = (r[0], str(r[9]), str(r[10]))
        if key in seen1:
            continue
        seen1.add(key)
        afterhours = is_afterhours(r[17])
        severity = 'CRITICAL' if (pct >= 50 or afterhours) else 'HIGH'
        s1.append({
            'Reservation No': r[0], 'Room No': r[6],
            'Guest Name': str(r[12])[:40] if r[12] else '-',
            'Status': r[2], 'Arrival': r[3], 'Departure': r[4],
            'Rate Lama': r2, 'Rate Baru': r1, 'Selisih': v,
            'Penurunan %': f"{pct:.1f}%",
            'Staff ID (ubah)': str(r[14]).strip() if r[14] else 'TIDAK ADA',
            'Staff ID (approve)': str(r[15]).strip() if r[15] else 'TIDAK ADA',
            'Change Date': r[16], 'Change Time': r[17],
            'Dini Hari?': 'YA ⚠️' if afterhours else 'Tidak',
            'Severity': severity,
            'Keterangan': f"Rate turun {pct:.0f}%{' — DINI HARI!' if afterhours else ''}",
        })
    s1.sort(key=lambda x: float(x['Penurunan %'].replace('%', '')), reverse=True)

    # Sheet 2: Rate hampir nol
    s2, seen2 = [], set()
    for r in rows:
        r1 = parse_amount(r[9])
        r2 = parse_amount(r[10])
        if not (r1 and r2 and r1 < 10000 and r2 > 50000):
            continue
        key = (r[0], str(r[9]), str(r[10]))
        if key in seen2:
            continue
        seen2.add(key)
        s2.append({
            'Reservation No': r[0], 'Room No': r[6],
            'Guest Name': str(r[12])[:40] if r[12] else '-',
            'Status': r[2], 'Rate Lama': r2, 'Rate Baru': r1,
            'Selisih': parse_amount(r[11]),
            'Staff ID': str(r[14]).strip() if r[14] else 'TIDAK ADA',
            'Change Date': r[16], 'Change Time': r[17],
            'Severity': 'CRITICAL',
            'Keterangan': f"Rate diturunkan ke Rp {r1:,.0f} — hampir nol!",
        })

    # Sheet 3: Dini hari
    s3, seen3 = [], set()
    for r in rows:
        if not is_afterhours(r[17]):
            continue
        v = parse_amount(r[11])
        if v is None or v == 0:
            continue
        key = (r[0], str(r[9]), str(r[10]), str(r[17]))
        if key in seen3:
            continue
        seen3.add(key)
        r1 = parse_amount(r[9])
        r2 = parse_amount(r[10])
        direction = 'TURUN ⬇️' if v < 0 else 'NAIK ⬆️'
        s3.append({
            'Reservation No': r[0], 'Room No': r[6],
            'Guest Name': str(r[12])[:40] if r[12] else '-',
            'Status': r[2], 'Rate Lama': r2, 'Rate Baru': r1, 'Selisih': v,
            'Arah': direction,
            'Staff ID': str(r[14]).strip() if r[14] else 'TIDAK ADA',
            'Change Date': r[16], 'Change Time': r[17],
            'Severity': 'CRITICAL' if v < 0 else 'HIGH',
            'Keterangan': f"Rate diubah pukul {r[17]} ({direction})",
        })

    # Sheet 4: Tanpa staff ID
    s4, seen4 = [], set()
    for r in rows:
        id1 = str(r[14]).strip() if r[14] else None
        if id1 and id1 not in ('None', '?', '', 'nan'):
            continue
        v = parse_amount(r[11])
        if v is None or v == 0:
            continue
        key = (r[0], str(r[9]), str(r[10]))
        if key in seen4:
            continue
        seen4.add(key)
        s4.append({
            'Reservation No': r[0], 'Room No': r[6],
            'Guest Name': str(r[12])[:40] if r[12] else '-',
            'Status': r[2],
            'Rate Lama': parse_amount(r[10]), 'Rate Baru': parse_amount(r[9]),
            'Selisih': v, 'Change Date': r[16], 'Change Time': r[17],
            'Severity': 'HIGH',
            'Keterangan': 'Rate berubah tanpa ID staff tercatat',
        })

    # Sheet 5: Multiple changes per reservasi
    resv_day = {}
    for r in rows:
        key = (r[0], r[16])
        resv_day.setdefault(key, []).append(r)
    s5 = []
    for (resv, date), grp in resv_day.items():
        unique = list({(parse_amount(x[9]), parse_amount(x[10])): x for x in grp}.values())
        if len(unique) <= 1:
            continue
        ids = set(str(x[14]).strip() if x[14] else 'TIDAK ADA' for x in grp)
        rates = ' → '.join(
            f"Rp {parse_amount(x[10]):,.0f}→{parse_amount(x[9]):,.0f}"
            for x in unique if parse_amount(x[9]) and parse_amount(x[10])
        )
        s5.append({
            'Reservation No': resv, 'Change Date': date,
            'Room No': grp[0][6],
            'Guest Name': str(grp[0][12])[:40] if grp[0][12] else '-',
            'Status': grp[0][2], 'Jumlah Perubahan': len(unique),
            'Kronologi Rate': rates[:200],
            'Staff IDs': ', '.join(ids),
            'Times': ', '.join(str(x[17]) for x in unique),
            'Severity': 'CRITICAL' if len(unique) >= 5 else 'HIGH',
            'Keterangan': f"{len(unique)}x rate change dalam 1 hari",
        })
    s5.sort(key=lambda x: -x['Jumlah Perubahan'])

    return s1, s2, s3, s4, s5
