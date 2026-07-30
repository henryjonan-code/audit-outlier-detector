"""Excel report writer for Hotel Anomaly Detector"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HDR_FONT  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
HDR_FILL  = PatternFill('solid', fgColor='2E4057')
CRIT_FILL = PatternFill('solid', fgColor='FF4444')
HIGH_FILL = PatternFill('solid', fgColor='FF8C00')
MED_FILL  = PatternFill('solid', fgColor='FFD700')
OK_FILL   = PatternFill('solid', fgColor='90EE90')
NORM_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
THIN_BDR  = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
)
SEV_FILLS = {'CRITICAL': CRIT_FILL, 'HIGH': HIGH_FILL, 'MEDIUM': MED_FILL}
SEV_FONTS = {
    'CRITICAL': Font(name='Arial', bold=True, size=10, color='FFFFFF'),
    'HIGH': Font(name='Arial', bold=True, size=10),
    'MEDIUM': Font(name='Arial', bold=True, size=10),
}


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BDR


def write_data_row(ws, row_idx, values, sev=None):
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        if isinstance(val, float) and abs(val) >= 1:
            cell.value = f"{val:,.0f}"
        elif val is None:
            cell.value = '-'
        else:
            cell.value = val
        cell.font = NORM_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BDR
    # Color the severity column
    if sev:
        for c_idx, val in enumerate(values, 1):
            if str(val) in SEV_FILLS:
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.fill = SEV_FILLS[sev]
                cell.font = SEV_FONTS.get(sev, NORM_FONT)


def write_sheet(wb, name, headers, data_rows, col_widths, sev_key='Severity'):
    ws = wb.create_sheet(name[:31])
    style_header(ws, 1, len(headers))
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    if not data_rows:
        ws.cell(row=2, column=1, value='✅ Tidak ada temuan pada kategori ini')
        ws.cell(row=2, column=1).font = Font(name='Arial', italic=True, size=10, color='009900')
        return ws

    for r_idx, row in enumerate(data_rows, 2):
        sev = row.get(sev_key) if isinstance(row, dict) else None
        vals = [row.get(h, '-') if isinstance(row, dict) else row[i]
                for i, h in enumerate(headers)]
        write_data_row(ws, r_idx, vals, sev)
        # Color severity cell
        if sev and sev in SEV_FILLS:
            try:
                sev_col = headers.index(sev_key) + 1
                cell = ws.cell(row=r_idx, column=sev_col)
                cell.fill = SEV_FILLS[sev]
                cell.font = SEV_FONTS.get(sev, NORM_FONT)
            except ValueError:
                pass

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    return ws


def write_summary(wb, label, sheets_info):
    ws = wb.create_sheet('Ringkasan', 0)
    ws.sheet_properties.tabColor = '2E4057'

    ws['A1'] = f'HOTEL FO ANOMALY REPORT — {label}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2E4057')
    ws.merge_cells('A1:F1')
    ws['A2'] = f'Dibuat: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
    ws.merge_cells('A2:F2')

    headers = ['No', 'Sheet', 'Keterangan', 'Jumlah Temuan', 'Severity', 'Prioritas']
    style_header(ws, 4, len(headers))
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)

    for r_idx, (no, sheet_name, desc, count, sev, priority) in enumerate(sheets_info, 5):
        row_vals = [no, sheet_name, desc, count, sev, priority]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = NORM_FONT
            cell.border = THIN_BDR
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        if sev in SEV_FILLS:
            sev_cell = ws.cell(row=r_idx, column=5)
            sev_cell.fill = SEV_FILLS[sev]
            sev_cell.font = SEV_FONTS.get(sev, NORM_FONT)

    for col, w in zip('ABCDEF', [5, 32, 55, 16, 12, 20]):
        ws.column_dimensions[col].width = w
    return ws
